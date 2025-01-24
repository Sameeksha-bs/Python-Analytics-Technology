"""
<Module 2 - Assignment>


Copyright (c) 2021 -- This is the 2021 Spring B version of the Template
Licensed
Written by <> <---- Sameeksha Santhosh

# you can also rely on the docstring documentation from pandas on how to format dosctrings:
# https://pandas.pydata.org/pandas-docs/stable/development/contributing_docstring.html

"""

           
import pandas as pd
import json
import openpyxl as op


def import_csv(file_path):
      # Check the file extension and read the file accordingly
    if file_path.endswith(".csv"):
        try:
            df = pd.read_csv(file_path, header=0, index_col=False)
            return df
        except FileNotFoundError:
            print(f"File not found at path: {file_path}")
        except pd.errors.EmptyDataError:
            print(f"File at path: {file_path} is empty")
        except pd.errors.ParserError as e:
            print(f"Error occurred while parsing CSV file: {e}")
        except Exception as e:
            print(f"Error occurred while reading CSV file: {e}")
    return none


def import_text(file_path):
      if file_path.endswith(".txt"):
        try:
            df = pd.read_csv(file_path, delimiter="|")
            return df
        except FileNotFoundError:
            print(f"File not found at path: {file_path}")
        except pd.errors.EmptyDataError:
            print(f"File at path: {file_path} is empty")
        except pd.errors.ParserError as e:
            print(f"Error occurred while parsing text file: {e}")
        except Exception as e:
            print(f"Error occurred while reading text file: {e}")
        return none


def import_json(file_path):
      if file_path.endswith(".json"):
        try:
            data = pd.read_json(file_path)
            return data
        except FileNotFoundError:
            print(f"File not found at path: {file_path}")
        except pd.errors.JSONDecodeError as e:
            print(f"Error occurred while parsing JSON file: {e}")
        except Exception as e:
            print(f"Error occurred while reading JSON file: {e}")
        return none


def import_excel(file_path):
    if file_path.endswith(".xlsx"):
        try:
            wb = op.load_workbook(file_path)
        except FileNotFoundError:
            print("File not found.")
            return
        except Exception as e:
            print(e)
            return

        sheets = wb.sheetnames
        for sheet_name in sheets:
            # Loop through each sheet in the Excel file
            sheet = wb[sheet_name]
            if sheet.max_row > 1 and sheet.max_column > 1:
                # Loop through each row in the sheet
                for i in range(1, sheet.max_row+1):
                    row_values = []
                    for j in range(1, sheet.max_column+1):
                        cell_value = sheet.cell(row=i, column=j).value
                        if cell_value is not None:
                            row_values.append(cell_value)
                    if row_values:
                        # Create a DataFrame with the row values and print it
                        df = pd.DataFrame([row_values], columns=row_values)
                        df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=i-1, header=None)
                        print(f"\nDisplaying EXCEL File : \n\nData from Sheet: {sheet_name}\n")
                        print(df)
                        break
    else:
        print("Invalid file type. Please provide a CSV, TXT, JSON, or XLSX file.")
        return None


if __name__ == "__main__":
    # File paths
    csv_path = "/Users/sameekshabs/Downloads/Neural_data.csv"
    text_path = "/Users/sameekshabs/Downloads/network_data.txt"
    json_path = "/Users/sameekshabs/Downloads/nested_data.json"
    excel_path = "/Users/sameekshabs/Downloads/Excel_report.xlsx"

    # Import and display data
    csv_data = import_csv(csv_path)
    print("\nDisplaying CSV Data:\n")
    if csv_data is not None:
        print(csv_data)

    text_data = import_text(text_path)
    print("\nDisplaying Text Data:\n")
    if text_data is not None:
        for line in text_data:
            print(line, end='')

    json_data = import_json(json_path)
    print("\nDisplaying JSON Data:\n")
    if json_data is not None:
        print(json_data)

    excel_data = import_excel(excel_path)
    print("\nDisplaying Excel Data:\n")
    if excel_data is not None:
        print(excel_data)
	